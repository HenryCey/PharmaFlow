"""
Shared report export utilities — Sprint 6.

Per the Sprint 6 brief: "Use reusable export utilities where appropriate.
If a shared export service does not yet exist, create one that can be
reused by future modules." Every report view hands this module the same
plain shape — a title, a list of column headers, and a list of rows of
plain (already-stringified/numeric) values, never HTML — so CSV, XLSX
and PDF stay three renderers of one shared, format-agnostic dataset
instead of three independent implementations. Any future module that
needs to export a table can call export_response() the same way.

Deliberately NOT reused for report *display* tables: those go through
components/_table.html and carry rendered HTML (badges, links) that
would be meaningless in a spreadsheet or PDF — each BaseReportView
subclass supplies get_export_rows() as the plain-value counterpart of
its display rows precisely so this module never has to sniff HTML out
of a cell.
"""
import csv
from datetime import datetime

from django.http import HttpResponse
from django.utils import timezone

SUPPORTED_FORMATS = ("csv", "xlsx", "pdf")


def _filename(base_name, extension):
    stamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_{stamp}.{extension}"


def _stringify_row(row):
    return ["" if value is None else str(value) for value in row]


def export_csv(*, base_name, title, headers, rows):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{_filename(base_name, "csv")}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(_stringify_row(row))
    return response


def _xlsx_safe_value(value):
    """
    openpyxl refuses to write a timezone-aware datetime (Excel's own
    file format has no concept of a UTC offset on a datetime cell) —
    Sprint 6 TESTBUILD v1 QA follow-up: every report built on
    sales_trend()/purchases_by_date() (Daily/Weekly/Monthly Sales,
    Revenue, Purchase Cost Analysis) annotates `period` via Django's
    Trunc*() on a DateTimeField, which returns a tz-aware datetime
    when USE_TZ=True — this crashed XLSX export on exactly those report
    pages. Converted to the pharmacy's local wall-clock time, then
    stripped of tzinfo, rather than left as UTC-with-no-marker or
    turned into a plain string, so the exported figure still reads as
    the same date/time shown on screen.
    """
    if isinstance(value, datetime) and value.tzinfo is not None:
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def export_xlsx(*, base_name, title, headers, rows):
    """
    openpyxl is used here only — it isn't a project-wide dependency
    otherwise, so it's imported lazily rather than at module load time,
    keeping every other view's import cost unchanged if xlsx export is
    never actually requested on a given request. It IS listed in
    requirements.txt; if this import fails, the environment simply
    hasn't installed it (`pip install -r requirements.txt`) — see
    Sprint 6 TESTBUILD v1 QA Issue 1.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "Excel export requires the 'openpyxl' package, which is listed in "
            "requirements.txt but not installed in this environment. Run "
            "`pip install -r requirements.txt` (or `pip install openpyxl`) and retry."
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (title or "Report")[:31]  # Excel's own 31-char sheet-name limit

    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)

    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=_xlsx_safe_value(value))

    for col_index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{_filename(base_name, "xlsx")}"'
    workbook.save(response)
    return response


def export_pdf(*, base_name, title, headers, rows, subtitle=None):
    """
    Sprint 6 TESTBUILD v1 QA Issue 2: WeasyPrint requires GTK/GObject
    native libraries (libgobject-2.0-0 etc.) that aren't present on a
    stock Windows install and are painful to provision there. ReportLab
    is a pure-Python PDF library — no native/system dependency at all —
    so this trades WeasyPrint's HTML+CSS templating for building the
    same visual layout (title, subtitle, generated timestamp, a banded
    header/zebra-striped table) directly with ReportLab's Platypus
    layout API. The previous HTML template
    (reports/exports/_pdf_report.html) is removed since nothing renders
    it anymore.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise ImportError(
            "PDF export requires the 'reportlab' package, which is listed in "
            "requirements.txt but not installed in this environment. Run "
            "`pip install -r requirements.txt` (or `pip install reportlab`) and retry."
        ) from exc
    from io import BytesIO

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        title=title,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PFTitle", parent=styles["Heading1"], fontSize=16, spaceAfter=2,
        textColor=colors.HexColor("#1e293b"),
    )
    subtitle_style = ParagraphStyle(
        "PFSubtitle", parent=styles["Normal"], fontSize=10, spaceAfter=2,
        textColor=colors.HexColor("#64748b"),
    )
    meta_style = ParagraphStyle(
        "PFMeta", parent=styles["Normal"], fontSize=8, spaceAfter=10,
        textColor=colors.HexColor("#94a3b8"),
    )
    header_cell_style = ParagraphStyle(
        "PFHeaderCell", parent=styles["Normal"], fontSize=8, leading=10,
        fontName="Helvetica-Bold", textColor=colors.HexColor("#334155"),
    )
    body_cell_style = ParagraphStyle(
        "PFBodyCell", parent=styles["Normal"], fontSize=8, leading=10,
        textColor=colors.HexColor("#1e293b"),
    )

    generated_at = timezone.localtime().strftime("%d %b %Y, %H:%M")
    elements = [Paragraph(f"PharmaFlow — {title}", title_style)]
    if subtitle:
        elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Paragraph(f"Generated {generated_at}", meta_style))
    elements.append(Spacer(1, 6))

    table_data = [[Paragraph(str(header), header_cell_style) for header in headers]]
    if rows:
        for row in rows:
            table_data.append([
                Paragraph("" if value is None else str(value), body_cell_style) for value in row
            ])
    else:
        table_data.append([Paragraph("No records found.", body_cell_style)] + [""] * (len(headers) - 1))

    column_width = doc.width / max(len(headers), 1)
    table = Table(table_data, colWidths=[column_width] * len(headers), repeatRows=1, hAlign="LEFT")

    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    if not rows:
        style_commands.append(("SPAN", (0, 1), (-1, 1)))
    for row_index in range(2, len(table_data), 2):  # zebra-stripe even body rows
        style_commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#f8fafc")))
    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_filename(base_name, "pdf")}"'
    return response


def export_response(*, export_format, base_name, title, headers, rows, subtitle=None):
    """Single entry point every report view calls — dispatches on
    `export_format` so views never import a specific renderer directly."""
    if export_format == "csv":
        return export_csv(base_name=base_name, title=title, headers=headers, rows=rows)
    if export_format == "xlsx":
        return export_xlsx(base_name=base_name, title=title, headers=headers, rows=rows)
    if export_format == "pdf":
        return export_pdf(base_name=base_name, title=title, headers=headers, rows=rows, subtitle=subtitle)
    raise ValueError(f"Unsupported export format: {export_format!r}")
