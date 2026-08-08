"""
Sprint 6 — Reports module view tests.

Renders real pages through the real URL/view/template stack (like
test_services.py's own test_dashboard_renders_for_all_three_roles_...),
so template bugs, permission-mixin misconfiguration, and export
plumbing are all caught here, not just the underlying service-layer math
(already covered by test_services.py).
"""
import csv
import io
from decimal import Decimal

import openpyxl
import pytest
from django.urls import reverse


def _get(client, username, url_name, **params):
    client.login(username=username, password="testpass123")
    response = client.get(reverse(f"reports:{url_name}"), params)
    client.logout()
    return response


# ---------------------------------------------------------------------------
# Permissions
# ---------------------------------------------------------------------------

def test_cashier_is_denied_every_report_category(seeded_pharmacy, client):
    """Cashier holds none of the four reports.* permissions per
    seed_role_permissions — every report URL, and the Reports Home page
    itself, must refuse them."""
    for url_name in ["home", "current_stock", "daily_sales", "purchase_history", "revenue"]:
        response = _get(client, "cashier1", url_name)
        assert response.status_code == 403, url_name


def test_owner_can_reach_every_report_page(seeded_pharmacy, client):
    """Owner holds all permissions (seed_role_permissions grants Owner
    everything) — every one of Sprint 6's report pages must render."""
    url_names = [
        "home",
        "current_stock", "low_stock", "near_expiry", "expired_stock",
        "stock_adjustments", "inventory_movements", "inventory_valuation",
        "daily_sales", "weekly_sales", "monthly_sales", "sales_date_range",
        "sales_by_drug", "sales_by_customer", "sales_by_cashier", "payment_method_summary",
        "purchase_history", "purchases_by_supplier", "purchases_by_drug",
        "purchase_cost_analysis", "outstanding_purchase_orders",
        "revenue", "financial_purchase_cost", "estimated_gross_profit",
        "inventory_value", "average_daily_sales",
    ]
    for url_name in url_names:
        response = _get(client, "owner1", url_name)
        assert response.status_code == 200, url_name


def test_reports_home_only_lists_categories_the_role_can_open(seeded_pharmacy, client):
    """Pharmacist (per seed_role_permissions's PHARMACIST_PERMS) holds
    view_inventory_reports and view_sales_reports but not
    view_purchase_reports or view_financial_reports — Reports Home must
    reflect exactly that, not show every category to everyone."""
    pharmacist = seeded_pharmacy["pharmacist"]
    assert pharmacist.has_perm("reports.view_inventory_reports") is True
    assert pharmacist.has_perm("reports.view_sales_reports") is True
    assert pharmacist.has_perm("reports.view_purchase_reports") is False
    assert pharmacist.has_perm("reports.view_financial_reports") is False

    response = _get(client, "pharmacist1", "home")
    assert response.status_code == 200
    content = response.content.decode()
    assert "Inventory Reports" in content
    assert "Sales Reports" in content
    assert "Purchase Reports" not in content
    assert "Financial Reports" not in content


# ---------------------------------------------------------------------------
# Report data correctness
# ---------------------------------------------------------------------------

def test_low_stock_report_shows_only_vitamin_c(seeded_pharmacy, client):
    response = _get(client, "owner1", "low_stock")
    content = response.content.decode()
    assert "Vitamin C 1000mg" in content
    assert "Paracetamol 500mg" not in content
    assert "Amoxicillin 250mg" not in content


def test_near_expiry_report_respects_days_filter(seeded_pharmacy, client):
    within_30 = _get(client, "owner1", "near_expiry", days=30).content.decode()
    assert "Paracetamol 500mg" in within_30
    assert "Amoxicillin 250mg" not in within_30

    within_500 = _get(client, "owner1", "near_expiry", days=500).content.decode()
    assert "Amoxicillin 250mg" in within_500


def test_sales_by_drug_report_shows_correct_revenue(seeded_pharmacy, client):
    response = _get(client, "owner1", "sales_by_drug")
    content = response.content.decode()
    # 5 sales x 2 units x ₦100 = ₦1,000.00 for Paracetamol (see
    # test_services.py's equivalent service-level assertion).
    assert "Paracetamol 500mg" in content
    assert "1,000.00" in content


def test_sales_by_customer_groups_walkins(seeded_pharmacy, client):
    response = _get(client, "owner1", "sales_by_customer")
    content = response.content.decode()
    assert "Jane Doe" in content
    assert "Walk-in" in content


def test_payment_method_summary_totals_match_sales_summary(seeded_pharmacy, client):
    response = _get(client, "owner1", "payment_method_summary")
    content = response.content.decode()
    assert "Cash" in content
    assert "1,500.00" in content  # 5 sales x ₦300 total each


def test_outstanding_purchase_orders_is_empty_once_received(seeded_pharmacy, client):
    """The fixture's one PurchaseOrder starts Draft and is received via
    receive_purchase() — once Received, it must drop off Outstanding."""
    response = _get(client, "owner1", "outstanding_purchase_orders")
    assert "No records found" in response.content.decode()


def test_purchase_history_status_filter(seeded_pharmacy, client):
    from apps.purchases.models import STATUS_RECEIVED

    matching = _get(client, "owner1", "purchase_history", purchase_status=STATUS_RECEIVED)
    assert seeded_pharmacy["purchase_order"].purchase_number in matching.content.decode()

    non_matching = _get(client, "owner1", "purchase_history", purchase_status="draft")
    assert "No records found" in non_matching.content.decode()


def test_estimated_gross_profit_matches_service_layer(seeded_pharmacy, client):
    from apps.reports.services import financial_report_service as fin

    expected = fin.estimated_gross_profit()
    response = _get(client, "owner1", "estimated_gross_profit")
    content = response.content.decode()
    assert f"{expected['revenue']:,.2f}" in content
    assert f"{expected['estimated_gross_profit']:,.2f}" in content


# ---------------------------------------------------------------------------
# Filters
# ---------------------------------------------------------------------------

def test_date_to_before_date_from_shows_validation_error(seeded_pharmacy, client):
    response = _get(client, "owner1", "sales_date_range", date_from="2026-08-10", date_to="2026-08-01")
    assert response.status_code == 200
    assert "Date To must be on or after Date From" in response.content.decode()


def test_drug_filter_narrows_current_stock_report(seeded_pharmacy, client):
    vitamin_c = seeded_pharmacy["vitamin_c"]
    response = _get(client, "owner1", "current_stock", drug=vitamin_c.pk)
    content = response.content.decode()
    # Scoped to the table itself, not the whole page — the filter form's
    # own <select name="drug"> legitimately lists every drug as an
    # option regardless of which one is currently selected.
    table_html = content.split('id="print-content"')[1]
    assert "Vitamin C 1000mg" in table_html
    assert "Paracetamol 500mg" not in table_html


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------

def test_csv_export_contains_expected_headers_and_data(seeded_pharmacy, client):
    response = _get(client, "owner1", "current_stock", export="csv")
    assert response.status_code == 200
    assert response["Content-Type"] == "text/csv"
    assert "attachment" in response["Content-Disposition"]

    rows = list(csv.reader(response.content.decode().splitlines()))
    assert rows[0] == ["Drug", "Category", "Unit", "Cost Price", "Selling Price", "Current Stock", "Reorder Level", "Status"]
    drug_names = {row[0] for row in rows[1:]}
    assert "Paracetamol 500mg" in drug_names
    # Export rows are plain values, not currency-symbol-prefixed display
    # strings — confirms get_export_rows() stayed independent of get_table().
    assert all(not row[3].startswith("\u20a6") for row in rows[1:])


def test_xlsx_export_is_a_valid_workbook_with_expected_rows(seeded_pharmacy, client):
    response = _get(client, "owner1", "sales_by_drug", export="xlsx")
    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header_row = [cell.value for cell in sheet[1]]
    assert header_row == ["Drug", "Quantity Sold", "Revenue"]
    drug_names = [row[0].value for row in sheet.iter_rows(min_row=2)]
    assert "Paracetamol 500mg" in drug_names


def test_xlsx_export_handles_timezone_aware_period_values(seeded_pharmacy, client):
    """Regression test: Daily/Weekly/Monthly Sales, Revenue, and Purchase
    Cost Analysis all annotate `period` via Django's Trunc*() on a
    DateTimeField, which returns a tz-aware datetime under USE_TZ=True.
    openpyxl raises TypeError on a tz-aware datetime cell — this crashed
    XLSX export on every trend-based report page (caught during Sprint 6
    TESTBUILD v1->v2 QA verification, not covered by the original test
    suite since sales_by_drug/sales_by_customer don't carry a `period`
    column)."""
    for url_name in ["daily_sales", "weekly_sales", "monthly_sales", "revenue", "purchase_cost_analysis"]:
        response = _get(client, "owner1", url_name, export="xlsx")
        assert response.status_code == 200, url_name
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        assert workbook.active["A1"].value is not None


def test_pdf_export_returns_a_valid_pdf(seeded_pharmacy, client):
    """ReportLab is a pure-Python dependency (Sprint 6 TESTBUILD v1 QA
    Issue 2 — replaced WeasyPrint, which required native GTK/GObject
    libraries unavailable on a stock Windows install). No
    importorskip needed: unlike WeasyPrint, there's no system library
    that can be missing underneath it."""
    response = _get(client, "owner1", "purchase_history", export="pdf")
    assert response.status_code == 200
    assert response["Content-Type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")
    # A real, non-trivial document — not just an empty/blank PDF shell.
    assert len(response.content) > 1500


def test_export_urls_preserve_active_filters(seeded_pharmacy, client):
    """Sprint 6 brief: exports should reflect what's on screen — the
    export link for a filtered view must carry the same filter, not the
    unfiltered full report."""
    vitamin_c = seeded_pharmacy["vitamin_c"]
    client.login(username="owner1", password="testpass123")
    response = client.get(reverse("reports:current_stock"), {"drug": vitamin_c.pk})
    client.logout()
    assert f"drug={vitamin_c.pk}" in response.context["export_urls"]["csv"]


# ---------------------------------------------------------------------------
# Pagination (UI Contract: "Every table should support ... Pagination")
# ---------------------------------------------------------------------------

def test_inventory_movement_report_paginates_past_25_rows(seeded_pharmacy, client, django_user_model):
    from apps.stock.models import MOVEMENT_ADJUSTMENT
    from apps.stock.services import record_movement

    paracetamol = seeded_pharmacy["paracetamol"]
    owner = seeded_pharmacy["owner"]
    # The fixture already wrote 10 sale movements; add enough more that
    # the combined total clears the 25-per-page threshold.
    for _ in range(20):
        record_movement(drug=paracetamol, movement_type=MOVEMENT_ADJUSTMENT, quantity=Decimal("1"), user=owner)

    page_one = _get(client, "owner1", "inventory_movements")
    assert page_one.status_code == 200
    assert page_one.context["page_obj"].paginator.num_pages >= 2

    page_two = _get(client, "owner1", "inventory_movements", page=2)
    assert page_two.status_code == 200
    assert page_two.context["page_obj"].number == 2


# ---------------------------------------------------------------------------
# Button consistency (Sprint 6 TESTBUILD v1 QA Issue 3)
# ---------------------------------------------------------------------------

def test_action_buttons_share_identical_sizing_classes(seeded_pharmacy, client):
    """Export CSV/Excel/PDF, Print, Apply Filters, and Reset must all
    carry the exact same height/padding/typography classes
    (`px-4 py-2 text-sm font-medium`) — this is what QA flagged as
    inconsistent in TESTBUILD v1 (Apply Filters rendering larger than
    the rest). Parses the actual rendered tags rather than trusting the
    template source, so a future edit that silently drops a class would
    fail this test."""
    import re

    response = _get(client, "owner1", "current_stock")
    content = response.content.decode()

    sizing_classes = {"px-4", "py-2", "text-sm", "font-medium", "rounded-md", "inline-flex"}
    labels = ["Export CSV", "Export Excel", "Export PDF", "Print", "Apply Filters", "Reset"]

    for label in labels:
        # Matches the opening <a ...> or <button ...> tag that directly
        # precedes this label's text.
        match = re.search(r'<(a|button)\s+([^>]*class="([^"]*)"[^>]*)>\s*' + re.escape(label), content)
        assert match, f"Could not locate a rendered <a>/<button> for {label!r}"
        classes = set(match.group(3).split())
        missing = sizing_classes - classes
        assert not missing, f"{label!r} button is missing sizing classes {missing}"
